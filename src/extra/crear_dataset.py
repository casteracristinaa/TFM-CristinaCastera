"""
Genera DATASET-envenenado.xlsx recorriendo ../logs/5_videos_envenenados
Estructura esperada:
  ../logs/5_videos_envenenados/<ID>/<clip_XXX>/resultados.txt

Uso:
  python generar_dataset_envenenado.py

Requiere: openpyxl  (pip install openpyxl)
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = Path("../../logs/5_videos_envenenados")
OUTPUT   = Path("DATASET-envenenado.xlsx")


def parse_resultados(txt_path: Path) -> dict:
    data = {}
    with open(txt_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if ":" in line:
                key, val = line.split(":", 1)
                data[key.strip()] = val.strip()
    return data


def main():
    rows = []
    for id_folder in sorted(BASE_DIR.iterdir()):
        if not id_folder.is_dir():
            continue
        for clip_folder in sorted(id_folder.iterdir()):
            if not clip_folder.is_dir():
                continue
            txt_file = clip_folder / "resultados.txt"
            if not txt_file.exists():
                print(f"  [AVISO] No encontrado: {txt_file}")
                continue
            data = parse_resultados(txt_file)
            rows.append({
                "ID":                 id_folder.name,
                "CLIP":               clip_folder.name,
                "velocidad_inyeccion": data.get("velocidad_inyeccion_px_s", ""),
                "angulo_triangulo":    data.get("angulo_triangulo_deg", ""),
            })

    # ── Estilos ────────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset"

    header_font  = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", start_color="1F4E79")
    header_align = Alignment(horizontal="center", vertical="center")
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers    = ["ID", "CLIP", "velocidad_inyeccion", "angulo_triangulo"]
    col_widths = [16,   14,     24,                    20]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22

    fill_even  = PatternFill("solid", start_color="D6E4F0")
    fill_odd   = PatternFill("solid", start_color="FFFFFF")
    data_font  = Font(name="Arial", size=10)
    data_align = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(rows, 2):
        fill   = fill_even if row_idx % 2 == 0 else fill_odd
        values = [
            row["ID"],
            row["CLIP"],
            row["velocidad_inyeccion"],
            row["angulo_triangulo"],
        ]
        for col_idx, val in enumerate(values, 1):
            try:
                val = float(val)
            except (ValueError, TypeError):
                pass
            cell            = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font       = data_font
            cell.fill       = fill
            cell.alignment  = data_align
            cell.border     = border

    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"✓ Archivo guardado: {OUTPUT.resolve()}")
    print(f"✓ Total de clips procesados: {len(rows)}")


if __name__ == "__main__":
    main()