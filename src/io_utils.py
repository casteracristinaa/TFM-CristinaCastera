"""
IO UTILS MODULE
---------------
Se encarga de:

- Crear carpetas de salida
- Guardar frames específicos
- Guardar resultados en un único resultados.txt
"""

import os
import cv2
import config
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


# =================================================
# CREAR DIRECTORIO
# =================================================

def ensure_dir(path):
    """
    Crea un directorio si no existe.
    """
    os.makedirs(path, exist_ok=True)


# =================================================
# GUARDAR FRAME
# =================================================

def save_frame(video_path, frame_idx, output_path):
    """
    Guarda un frame específico del vídeo.
    """

    if frame_idx is None:
        return

    cap = cv2.VideoCapture(video_path)

    if not cap.isOpened():
        raise RuntimeError(f"No se pudo abrir el vídeo: {video_path}")

    cap.set(cv2.CAP_PROP_POS_FRAMES, frame_idx)
    ret, frame = cap.read()

    if ret:
        cv2.imwrite(output_path, frame)
        if config.DEBUG:
            print(f"[INFO] Frame guardado: {output_path}")
    else:
        print(f"[WARN] No se pudo leer el frame {frame_idx}")

    cap.release()






# =================================================
# VIDEO DEBUG RACKING
# =================================================
def save_debug_video(video_path, out_dir, events, trajectory):

    cap = cv2.VideoCapture(video_path)
    fps = cap.get(cv2.CAP_PROP_FPS)
    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    output_path = os.path.join(out_dir, "debug_tracking.mp4")

    writer = cv2.VideoWriter(
        output_path,
        cv2.VideoWriter_fourcc(*"mp4v"),
        fps,
        (w, h)
    )

    frame_idx = 0
    traj_dict = {t[0]: t for t in trajectory} if trajectory else {}

    while True:
        ret, frame = cap.read()
        if not ret:
            break


        # Dibujar tracking solo entre eventos
        if frame_idx in traj_dict:
            _, x, y, _, _= traj_dict[frame_idx]
            if x != -1:
                cv2.circle(frame, (x,y), 6, (0,0,255), -1)

        writer.write(frame)
        frame_idx += 1

    writer.release()
    cap.release()


# =================================================
# GUARDAR RESULTADOS UNIFICADOS
# =================================================

def save_results(out_dir, artifact_frame, fps, events, velocidad_inyeccion=None, angulo_deg=None, angulo_status=None, frame_min_angle=None, min_angle=None, duracion_total=None, tiempo_triangulo=None, std_angulo=None):
    """
    Guarda un único archivo resultados.txt
    con frames, segundos, velocidad de inyección y ángulo.
    """

    ensure_dir(out_dir)

    injection  = events.get("injection")
    interno = events.get("interno")


    result_path = os.path.join(out_dir, "resultados.txt")

    with open(result_path, "w") as f:

        # -------------------------
        # INYECCIÓN
        # -------------------------
        if injection is not None:
            f.write(f"frame_inyeccion: {injection}\n")
            f.write(f"segundo_inyeccion: {injection / fps:.3f}\n\n")
        else:
            f.write("frame_inyeccion: None\n")
            f.write("segundo_inyeccion: None\n\n")

        # -------------------------
        # INTERNO
        # -------------------------
        if interno is not None:
            f.write(f"frame_interno: {interno}\n")
            f.write(f"segundo_interno: {interno / fps:.3f}\n\n")
        else:
            f.write("frame_interno: None\n")
            f.write("segundo_interno: None\n\n")

        # -------------------------
        # ASPIRACION
        # -------------------------
        if artifact_frame is not None:
            f.write(f"frame_aspiracion: {artifact_frame}\n")
            f.write(f"segundo_aspiracion: {artifact_frame / fps:.3f}\n\n")
        else:
            f.write("frame_aspiracion: None\n")
            f.write("segundo_aspiracion: None\n\n")

        # -------------------------
        # VELOCIDAD DE INYECCIÓN
        # -------------------------
        if velocidad_inyeccion is not None:
            f.write(f"velocidad_inyeccion_px_s: {velocidad_inyeccion:.3f}\n\n")
        else:
            f.write("velocidad_inyeccion_px_s: None\n\n")

        # -------------------------
        # MENOR ÁNGULO DEL TRIÁNGULO
        # -------------------------
        if frame_min_angle is not None:
            f.write(f"frame_min_angle: {frame_min_angle}\n")
        else:
            f.write("frame_min_angle: None\n")

        # -------------------------
        # ÁNGULO DEL TRIÁNGULO
        # -------------------------
        if angulo_deg is not None and angulo_status == "OK":
            f.write(f"angulo_triangulo_deg: {angulo_deg:.2f}\n")
        elif angulo_status is not None and angulo_status != "OK":
            f.write(f"angulo_triangulo_status: {angulo_status}\n")
        else:
            f.write("angulo_triangulo_deg: None\n")

        f.write("\n")

        if min_angle is not None:
            f.write(f"min_angle: {min_angle:.2f}\n")
        else:
            f.write("min_angle: None\n")


        # -------------------------
        # MÉTRICAS DE LA SEÑAL
        # -------------------------
        if duracion_total is not None:
            f.write(f"\nduracion_total_s: {duracion_total:.3f}\n")
        else:
            f.write("\nduracion_total_s: None\n")

        if tiempo_triangulo is not None:
            f.write(f"tiempo_triangulo_s: {tiempo_triangulo:.3f}\n")
        else:
            f.write("tiempo_triangulo_s: None\n")

        if std_angulo is not None:
            f.write(f"std_angulo: {std_angulo:.3f}\n")
        else:
            f.write("std_angulo: None\n")
       

    if config.DEBUG:
        print(f"[INFO] Resultados guardados en: {result_path}")


# =================================================
# VIDEO DEBUG CON OVERLAYS DE TEXTO
# =================================================

def create_debug_video_with_overlays(video_path, output_dir, events, artifact_frame, fps):
    """
    Crea un video debug con overlays de texto en frames clave durante 2 segundos.
    """
    cap = cv2.VideoCapture(video_path)
    if not cap.isOpened():
        raise RuntimeError(f"No se pudo abrir el vídeo: {video_path}")

    w = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    h = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))

    output_path = os.path.join(output_dir, "debug_overlays.mp4")

    writer = cv2.VideoWriter(
        output_path,
        cv2.VideoWriter_fourcc(*"mp4v"),
        fps,
        (w, h)
    )

    # Definir frames clave y textos
    overlays = {
        events.get("injection"): "frame injeccion",
        events.get("interno"): "frame interno",
        artifact_frame: "frame aspiracion",
        events.get("retroceso"): "frame retroceso"
    }

    # Duración en frames: 2 segundos
    duration_frames = int(fps * 2)

    frame_idx = 0
    while True:
        ret, frame = cap.read()
        if not ret:
            break

        # Verificar si este frame tiene overlay
        text = None
        for start_frame, txt in overlays.items():
            if start_frame is not None and start_frame <= frame_idx < start_frame + duration_frames:
                text = txt  # Sobrescribir si hay múltiples, mostrando el último

        if text:
            # Añadir texto en la parte superior izquierda
            cv2.putText(frame, text, (10, 50), cv2.FONT_HERSHEY_SIMPLEX, 1.5, (0, 0, 255), 3, cv2.LINE_AA)

        writer.write(frame)
        frame_idx += 1

    writer.release()
    cap.release()

    if config.DEBUG:
        print(f"[INFO] Video debug con overlays guardado en: {output_path}")


# =================================================
# DATASET GLOBAL (.xlsx)
# =================================================

DATASET_HEADERS = ["ID", "CLIP", "velocidad_inyeccion", "angulo_triangulo"]


def guardar_en_dataset(id_paciente, clip_name, velocidad_inyeccion, angulo_triangulo):
    """
    Añade (o actualiza) una fila en el dataset global config.DATASET_PATH
    con las métricas clave de un clip: ID, CLIP, velocidad_inyeccion y
    angulo_triangulo.

    Si ya existe una fila para ese (ID, CLIP) -p. ej. al relanzar el
    pipeline tras corregir un error- se sobreescribe en lugar de
    duplicarla.
    """
    if os.path.exists(config.DATASET_PATH):
        wb = load_workbook(config.DATASET_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Dataset"
        ws.append(DATASET_HEADERS)
        for col_idx in range(1, len(DATASET_HEADERS) + 1):
            ws.cell(row=1, column=col_idx).font = Font(bold=True)

    valores = [id_paciente, clip_name, velocidad_inyeccion, angulo_triangulo]

    fila_existente = None
    for row in ws.iter_rows(min_row=2, max_col=2):
        if row[0].value == id_paciente and row[1].value == clip_name:
            fila_existente = row[0].row
            break

    if fila_existente:
        for col_idx, valor in enumerate(valores, start=1):
            ws.cell(row=fila_existente, column=col_idx).value = valor
    else:
        ws.append(valores)

    wb.save(config.DATASET_PATH)

    if config.DEBUG:
        print(f"[INFO] Dataset actualizado: {config.DATASET_PATH} (ID={id_paciente}, CLIP={clip_name})")